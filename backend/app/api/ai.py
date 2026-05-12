from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import func, distinct
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime, date, timedelta
from ..database import get_db
from ..models.models import Task, User, Subject, PomodoroSession, ChatHistory, UserProfile, TaskChatHistory, Schedule
from ..utils.auth import get_current_active_user
import os
import io
import httpx
import json
import base64
from dotenv import load_dotenv
from ..utils.cache import invalidate_cache

# tải các biến môi trường
load_dotenv()

router = APIRouter(prefix="/stms/ai", tags=["ai"])

API_KEYS_RAW = os.getenv("OPENROUTER_API_KEYS", "")
API_KEYS = [k.strip() for k in API_KEYS_RAW.split(",") if k.strip()]
if not API_KEYS:
    single = os.getenv("OPENROUTER_API_KEY", "")
    if single:
        API_KEYS = [single]

AI_API_KEY = API_KEYS[0] if API_KEYS else ""
# Tạo danh sách các key
ACTIVE_KEYS_LIST = list(API_KEYS)

AI_URL = "https://openrouter.ai/api/v1/chat/completions"

# Các model miễn phí
MODELS = [
    "qwen/qwen3.6-plus-preview:free",
    "nvidia/nemotron-3-super-120b-a12b:free",
    "stepfun/step-3.5-flash:free",
    "google/gemma-3-27b-it:free",
    "z-ai/glm-4.5-air:free"
]

# Model hỗ trợ Vision (đọc ảnh) - miễn phí trên OpenRouter
VISION_MODELS = [
    "google/gemma-3-27b-it:free",
    "meta-llama/llama-4-scout:free",
    "qwen/qwen-2.5-vl-72b-instruct:free",
]


def get_user_context(db: Session, user: User) -> str:
    """Build a context string from the user's profile for AI personalization. Uses English labels to avoid language bias."""
    profile = db.query(UserProfile).filter(UserProfile.user_id == user.id).first()
    parts = []
    name = user.full_name or user.username
    parts.append(f"Name: {name}")
    if profile:
        if profile.grade_level:
            parts.append(f"Grade: {profile.grade_level}")
        if profile.school_name:
            parts.append(f"School: {profile.school_name}")
        if profile.study_goal:
            parts.append(f"Study goal: {profile.study_goal}")
        if profile.daily_study_target:
            parts.append(f"Daily task target: {profile.daily_study_target} tasks")
        if profile.gender:
            parts.append(f"Gender: {profile.gender}")
    # Đếm số môn học
    subjects = db.query(Subject).filter(Subject.user_id == user.id).all()
    if subjects:
        subject_names = [s.subject_name for s in subjects if getattr(s, 'subject_name', None)]
        if subject_names:
            parts.append(f"Subjects: {', '.join(subject_names)}")
    return "\n".join(parts)

import asyncio
import time as _time

async def _try_model(model_id: str, prompt, headers: dict, max_tokens: int) -> tuple:
    """Thử một model, trả về (model_id, content) hoặc (model_id, None). prompt có thể là str hoặc list (multimodal)"""
    try:
        # Hỗ trợ cả text thuần và multimodal content
        if isinstance(prompt, str):
            messages_content = prompt
        else:
            messages_content = prompt  # Đã là list [{"type": "text"}, {"type": "image_url"}]
        
        payload = {
            "model": model_id,
            "messages": [{"role": "user", "content": messages_content}],
            "max_tokens": max_tokens,
        }
        async with httpx.AsyncClient(timeout=45.0) as client:
            t0 = _time.time()
            response = await client.post(AI_URL, json=payload, headers=headers)
            elapsed = round(_time.time() - t0, 1)
            data = response.json()
            
            if "error" in data:
                err_msg = data["error"].get("message", "Unknown error")
                print(f"[AI] {model_id} error ({elapsed}s): {err_msg}")
                return (model_id, None, err_msg)
            
            if "choices" in data and data["choices"]:
                content = data["choices"][0].get("message", {}).get("content", "")
                if content:
                    print(f"[AI] {model_id} OK ({elapsed}s, {len(content)} chars)")
                    return (model_id, content, None)
            
            print(f"[AI] {model_id} empty response ({elapsed}s)")
            return (model_id, None, "empty response")
    except Exception as e:
        print(f"[AI] {model_id} exception: {e}")
        return (model_id, None, str(e))

async def call_ai_async(prompt: str, max_tokens: int = 4096) -> str:
    """Call OpenRouter API - loops through available keys and models."""
    global ACTIVE_KEYS_LIST
    
    if not API_KEYS:
        load_dotenv()
        temp_raw = os.getenv("OPENROUTER_API_KEYS", "")
        temp_keys = [k.strip() for k in temp_raw.split(",") if k.strip()]
        if not temp_keys:
            single = os.getenv("OPENROUTER_API_KEY", "")
            if single: temp_keys = [single]
        if not temp_keys:
            raise Exception("API Key chưa được cấu hình. Vui lòng thêm OPENROUTER_API_KEYS vào .env và khởi động lại Backend.")
        ACTIVE_KEYS_LIST = list(temp_keys)

    if not ACTIVE_KEYS_LIST:
        ACTIVE_KEYS_LIST = list(API_KEYS)
        
    last_key_error = ""
    for current_key in list(ACTIVE_KEYS_LIST):
        # Đảm bảo key hợp lệ
        if not current_key or len(current_key) < 10:
            continue
            
        headers = {
            "Authorization": f"Bearer {current_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "http://localhost:8080",
            "X-Title": "EduFlow",
        }
        
        errors = []
        key_exhausted = False
        
        # Chạy các model lần lượt
        for model_id in MODELS:
            _, content, err = await _try_model(model_id, prompt, headers, max_tokens)
            if content:
                # Trả về ngay nếu có Model trả lời thành công
                return content
                
            errors.append(f"{model_id}: {err}")
            
            # Nếu key hết hạn hoặc quá tải
            if err and any(code in err.lower() for code in ["401", "403", "402", "exceeded", "limit", "credits", "quota"]):
                key_exhausted = True
                break
                
        if key_exhausted:
            print(f"[AI] Key {current_key[:12]}... hết hạn/quá tải. Chuyển sang Key tiếp theo!")
            if current_key in ACTIVE_KEYS_LIST:
                ACTIVE_KEYS_LIST.remove(current_key)
            last_key_error = f"Key {current_key[:12]}... lỗi: {err}"
            continue # Thử key tiếp theo
            
        # Nếu chạy hết các AI cho Key này mà đều lỗi (nhưng không phải lỗi Token/Rate limit)
        raise Exception(f"Tất cả {len(MODELS)} AI models đều thất bại. Chi tiết lỗi: {', '.join(errors)}")
        
    raise Exception(f"Đã thử dùng TẤT CẢ các API Key nhưng đều hết hạn/bị chặn! Cần update Key mới. Lỗi cuối: {last_key_error}")

def call_ai(prompt: str, max_tokens: int = 4096) -> str:
    """Sync wrapper"""
    try:
        loop = asyncio.get_running_loop()
        import concurrent.futures
        with concurrent.futures.ThreadPoolExecutor() as pool:
            future = pool.submit(asyncio.run, call_ai_async(prompt, max_tokens))
            return future.result(timeout=120)
    except RuntimeError:
        # Không có loop đang chạy, chạy trực tiếp
        return asyncio.run(call_ai_async(prompt, max_tokens))


async def call_ai_vision_async(image_bytes: bytes, filename: str, question: str, max_tokens: int = 4096) -> str:
    """Gọi Vision AI để đọc/phân tích nội dung ảnh. Encode ảnh thành base64 và gửi qua multimodal API."""
    global ACTIVE_KEYS_LIST
    
    # Xác định MIME type
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else 'png'
    mime_map = {'png': 'image/png', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'gif': 'image/gif', 'webp': 'image/webp', 'bmp': 'image/bmp'}
    mime_type = mime_map.get(ext, 'image/png')
    
    # Encode ảnh thành base64
    b64_image = base64.b64encode(image_bytes).decode('utf-8')
    data_url = f"data:{mime_type};base64,{b64_image}"
    
    # Tạo multimodal content (chuẩn OpenAI Vision format)
    multimodal_content = [
        {"type": "text", "text": f"{question}\n\nHãy phân tích nội dung ảnh chi tiết. Nếu ảnh chứa văn bản, hãy trích xuất toàn bộ text. Nếu là bài tập, hãy giải. Trả lời bằng cùng ngôn ngữ với câu hỏi."},
        {"type": "image_url", "image_url": {"url": data_url}}
    ]
    
    keys_to_try = list(ACTIVE_KEYS_LIST) if ACTIVE_KEYS_LIST else list(API_KEYS)
    
    for current_key in keys_to_try:
        if not current_key or len(current_key) < 10:
            continue
        headers = {
            "Authorization": f"Bearer {current_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "http://localhost:8080",
            "X-Title": "EduFlow",
        }
        
        for model_id in VISION_MODELS:
            _, content, err = await _try_model(model_id, multimodal_content, headers, max_tokens)
            if content:
                print(f"[Vision AI] Quét ảnh '{filename}' thành công bằng {model_id}")
                return content
            if err and any(code in err.lower() for code in ["401", "403", "402", "exceeded", "limit", "credits", "quota"]):
                break
    
    raise Exception("Tất cả Vision models đều thất bại. Không thể quét ảnh.")

def extract_text_from_file(content: bytes, filename: str) -> str:
    """Trích xuất text từ nhiều định dạng file: PDF, Word, text, v.v."""
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    
    # PDF
    if ext == 'pdf':
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(content))
            pages_text = []
            for i, page in enumerate(reader.pages):
                text = page.extract_text()
                if text:
                    pages_text.append(f"--- Trang {i+1} ---\n{text}")
            if pages_text:
                return "\n\n".join(pages_text)
            return "[PDF không có text trích xuất được (có thể là ảnh scan)]"
        except Exception as e:
            return f"[Lỗi đọc PDF: {e}]"
    
    # Word (.docx)
    if ext in ('docx', 'doc'):
        try:
            from docx import Document
            doc = Document(io.BytesIO(content))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            if paragraphs:
                return "\n\n".join(paragraphs)
            return "[File Word trống hoặc không đọc được nội dung]"
        except Exception as e:
            return f"[Lỗi đọc Word: {e}]"
    
    # Excel (.xlsx)
    if ext in ('xlsx', 'xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            all_text = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                all_text.append(f"--- Sheet: {sheet} ---")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    all_text.append(" | ".join(cells))
            return "\n".join(all_text) if all_text else "[Excel trống]"
        except:
            return "[Không thể đọc file Excel, hãy thử cài openpyxl]"
    
    # Images - trả về bytes để xử lý Vision AI
    if ext in ('png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'):
        return "[IMAGE_FILE]"
    
    # File dạng text
    text_exts = ('txt', 'md', 'csv', 'json', 'py', 'js', 'java', 'html', 'css', 'xml', 'yaml', 'yml', 'sql', 'log', 'ts', 'jsx', 'tsx', 'c', 'cpp', 'h', 'rb', 'php', 'sh', 'bat', 'ini', 'cfg', 'env')
    if ext in text_exts:
        try:
            return content.decode('utf-8')
        except:
            try:
                return content.decode('latin-1')
            except:
                return "[Không thể đọc file text]"
    
    # Thử UTF-8 cho các đuôi mở rộng không xác định
    try:
        return content.decode('utf-8')
    except:
        return f"[Định dạng file .{ext} chưa được hỗ trợ. Hỗ trợ: PDF, Word, Excel, text, code, ảnh]"


class InsightResponse(BaseModel):
    insight: str

class ChatMessageResponse(BaseModel):
    id: int
    role: str
    message: str
    file_name: Optional[str] = None
    thread_id: Optional[str] = None
    created_at: datetime

    class Config:
        from_attributes = True

class ThreadResponse(BaseModel):
    thread_id: str
    title: str
    last_message: str
    created_at: datetime
    
# ---- Prompt instructions ----
NO_EMOJI_INSTRUCTION = """IMPORTANT FORMATTING RULES:
- Do NOT use emojis or emoticons in your response (no 📖🧑🎯🌐⭐🔑💡🚀 etc.)
- Use clean, professional Markdown formatting only (headings, bold, lists, tables, code blocks).
- Keep the tone academic and clear."""

LANGUAGE_DETECT_INSTRUCTION = """CRITICAL LANGUAGE RULE (MUST FOLLOW):
- Analyze the ACTUAL CONTENT of the document or material provided.
- You MUST respond in the EXACT SAME language as the document content. 
- (e.g., If the document is in Japanese, respond in Japanese. If French, respond in French. If Vietnamese, respond in Vietnamese).
- Do NOT randomly insert extraneous languages not present in the source material.
- If no document is provided, match the language of the task title."""

LANGUAGE_FOLLOW_USER_INSTRUCTION = """LANGUAGE RULE:
- You MUST respond in the EXACT SAME language the user is using in their latest message, UNLESS they explicitly ask you to reply in a specific language.
- (e.g., If the user asks in Korean, reply in Korean. If they type in Vietnamese but say 'giải thích bằng tiếng anh', you MUST reply in English).
- Do NOT hallucinate or insert random foreign characters/languages that the user did not use or ask for, unless explaining a technical term.
- This rule takes absolute priority over the document language."""

# ---- Sơ đồ phân tích AI ----
class TaskBreakdownRequest(BaseModel):
    goal: str
    
class SubTaskIdea(BaseModel):
    title: str
    estimated_pomodoros: int

class BreakdownResponse(BaseModel):
    subtasks: List[SubTaskIdea]

class StudyRequest(BaseModel):
    task_id: Optional[int] = None
    subject_id: Optional[int] = None

class StudyChatRequest(BaseModel):
    task_id: int
    message: str
    chat_type: str = "study" 

class ScheduleBlock(BaseModel):
    title: str
    description: str
    start_time: datetime
    end_time: datetime

class AutoScheduleResponse(BaseModel):
    blocks: List[ScheduleBlock]

# ---- AI ----
@router.get("/insights", response_model=InsightResponse)
async def get_study_insights(db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    if not AI_API_KEY:
        return {"insight": "Chưa cấu hình API Key."}
    try:
        tasks = db.query(Task).filter(Task.user_id == current_user.id).all()
        total_tasks = len(tasks)
        completed_tasks = len([t for t in tasks if t.status in ('completed', 'done')])
        pending_tasks = total_tasks - completed_tasks

        now = datetime.now()
        week_start = now - timedelta(days=7)
        focus_minutes = db.query(func.sum(PomodoroSession.total_focus_time))\
            .filter(PomodoroSession.user_id == current_user.id, PomodoroSession.session_date >= week_start)\
            .scalar() or 0
        focus_hours = round(float(focus_minutes) / 60, 1)

        user_ctx = get_user_context(db, current_user)
        prompt = f"""Bạn là EduFlow AI trợ lý học tập.
Thông tin học sinh:
{user_ctx}

Số liệu tuần qua:
- Hoàn thành: {completed_tasks}/{total_tasks} nhiệm vụ.
- Chờ xử lý: {pending_tasks}.
- Thời gian tập trung tuần qua: {focus_hours} giờ.
Viết 2-3 câu nhận xét ngắn gọn, thân thiện, động viên và gợi ý hành động cụ thể phù hợp với lớp và mục tiêu của học sinh."""

        result = await call_ai_async(prompt)
        return {"insight": result.strip()}
    except Exception as e:
        print(f"AI Error: {e}")
        return {"insight": "EduFlow AI đang nghỉ ngơi, vui lòng thử lại sau nhé."}

# ---- Chat with AI ----
@router.post("/chat")
async def chat_with_ai(
    message: str = Form(...),
    thread_id: str = Form("default"),
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if not AI_API_KEY:
        raise HTTPException(status_code=500, detail="API Key chưa được cấu hình")
    try:
        file_name = None
        if file:
            content = await file.read()
            file_name = file.filename
            text_content = extract_text_from_file(content, file.filename)
            
            if text_content == "[IMAGE_FILE]":
                # Sử dụng Vision AI để đọc ảnh
                try:
                    vision_reply = await call_ai_vision_async(content, file.filename, message or "Hãy mô tả và phân tích nội dung trong ảnh này.")
                    if vision_reply:
                        # Lưu tin nhắn user
                        user_msg = ChatHistory(user_id=current_user.id, role="user", message=message, file_name=file_name, thread_id=thread_id)
                        db.add(user_msg)
                        db.commit()
                        # Lưu tin nhắn AI
                        ai_msg = ChatHistory(user_id=current_user.id, role="ai", message=vision_reply, thread_id=thread_id)
                        db.add(ai_msg)
                        db.commit()
                        return {"reply": vision_reply}
                except Exception as ve:
                    print(f"[Vision AI] Fallback to text: {ve}")
                # Fallback nếu Vision không hoạt động
                prompt = f"Người dùng gửi ảnh '{file.filename}'. Đây là file ảnh nên tôi không đọc được nội dung text. Hãy cho biết bạn đã nhận ảnh và gợi ý gì với người dùng.\n\nCâu hỏi: {message}"
            else:
                prompt = f"Tôi có tài liệu '{file.filename}':\n\n{text_content[:12000]}\n\nCâu hỏi: {message}"
        else:
            prompt = message

        # Lưu tin nhắn user
        user_msg = ChatHistory(user_id=current_user.id, role="user", message=message, file_name=file_name, thread_id=thread_id)
        db.add(user_msg)
        db.commit()

        user_ctx = get_user_context(db, current_user)
        system = f"Bạn là EduFlow AI trợ lý học tập. Thông tin học sinh:\n{user_ctx}\n\nTrả lời bằng tiếng Việt, ngắn gọn, súc tích, phù hợp với trình độ và mục tiêu của học sinh."
        full_prompt = system + "\n\n" + prompt

        reply = await call_ai_async(full_prompt)

        # Lưu tin nhắn AI
        ai_msg = ChatHistory(user_id=current_user.id, role="ai", message=reply, thread_id=thread_id)
        db.add(ai_msg)
        db.commit()

        return {"reply": reply}
    except Exception as e:
        print(f"AI Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ---- Threads ----
@router.get("/threads")
async def get_threads(db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    """Lấy tất cả các luồng trò chuyện của user"""
    from sqlalchemy import desc
    threads_raw = db.query(
        ChatHistory.thread_id,
        func.min(ChatHistory.created_at).label("created_at"),
        func.max(ChatHistory.created_at).label("last_at")
    ).filter(
        ChatHistory.user_id == current_user.id
    ).group_by(ChatHistory.thread_id).order_by(desc("last_at")).all()

    result = []
    for t in threads_raw:
        first_user_msg = db.query(ChatHistory).filter(
            ChatHistory.user_id == current_user.id,
            ChatHistory.thread_id == t.thread_id,
            ChatHistory.role == "user"
        ).order_by(ChatHistory.created_at.asc()).first()
        
        title = first_user_msg.message[:40] if first_user_msg else "Cuộc trò chuyện mới"
        
        last_msg = db.query(ChatHistory).filter(
            ChatHistory.user_id == current_user.id,
            ChatHistory.thread_id == t.thread_id
        ).order_by(ChatHistory.created_at.desc()).first()

        result.append({
            "thread_id": t.thread_id,
            "title": title,
            "last_message": last_msg.message[:60] if last_msg else "",
            "created_at": str(t.created_at)
        })
    return result

# ---- Chat History ----
@router.get("/history")
async def get_chat_history(thread_id: str = "default", db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    messages = db.query(ChatHistory)\
        .filter(ChatHistory.user_id == current_user.id, ChatHistory.thread_id == thread_id)\
        .order_by(ChatHistory.created_at.asc())\
        .all()
    return [{"id": m.id, "role": m.role, "message": m.message, "file_name": m.file_name, "created_at": str(m.created_at)} for m in messages]

@router.delete("/history")
async def clear_all_history(db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    db.query(ChatHistory).filter(ChatHistory.user_id == current_user.id).delete()
    db.commit()
    return {"message": "Đã xóa toàn bộ lịch sử"}

@router.delete("/threads/{thread_id}")
async def delete_thread(thread_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    db.query(ChatHistory).filter(ChatHistory.user_id == current_user.id, ChatHistory.thread_id == thread_id).delete()
    db.commit()
    return {"message": "Đã xóa cuộc trò chuyện"}

# ---- Tải lên và phân tích tài liệu ----
@router.post("/upload-doc")
async def upload_and_analyze(
    file: UploadFile = File(...),
    question: str = Form("Hãy tóm tắt nội dung tài liệu này."),
    thread_id: str = Form("default"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if not AI_API_KEY:
        raise HTTPException(status_code=500, detail="API Key chưa cấu hình")
    
    # Lưu file upload
    upload_dir = f"uploads/ai/{current_user.id}"
    os.makedirs(upload_dir, exist_ok=True)
    file_path = f"{upload_dir}/{file.filename}"
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    # Trích xuất văn bản
    text_content = extract_text_from_file(content, file.filename)

    # Lưu tin nhắn
    user_msg = ChatHistory(user_id=current_user.id, role="user", message=question, file_name=file.filename, thread_id=thread_id)
    db.add(user_msg)
    db.commit()

    try:
        user_ctx = get_user_context(db, current_user)
        system = f"Bạn là EduFlow AI. Thông tin học sinh:\n{user_ctx}\n\nHọc sinh gửi tài liệu '{file.filename}'."
        
        if text_content == "[IMAGE_FILE]":
            # Sử dụng Vision AI để đọc ảnh
            try:
                vision_reply = await call_ai_vision_async(content, file.filename, question)
                if vision_reply:
                    ai_msg = ChatHistory(user_id=current_user.id, role="ai", message=vision_reply, thread_id=thread_id)
                    db.add(ai_msg)
                    db.commit()
                    return {"reply": vision_reply, "file_saved": file.filename}
            except Exception as ve:
                print(f"[Vision AI] Fallback to text: {ve}")
            # Fallback
            prompt = f"{system}\n\nĐây là file ảnh, tôi không đọc được nội dung text từ ảnh. Hãy cho biết đã nhận file và gợi ý.\n\nYêu cầu: {question}"
        else:
            prompt = f"{system}\n\nNội dung tài liệu:\n{text_content[:12000]}\n\nYêu cầu: {question}"
        
        reply = await call_ai_async(prompt)

        ai_msg = ChatHistory(user_id=current_user.id, role="ai", message=reply, thread_id=thread_id)
        db.add(ai_msg)
        db.commit()

        return {"reply": reply, "file_saved": file.filename}
    except Exception as e:
        print(f"AI Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ---- AI phân tích ----
@router.post("/breakdown-task", response_model=BreakdownResponse)
async def breakdown_task(
    goal: str = Form(...),
    file: Optional[UploadFile] = File(None),
    task_id: Optional[int] = Form(None),
    attachment_path: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if not AI_API_KEY:
        raise HTTPException(status_code=500, detail="API Key chưa cấu hình")
        
    user_ctx = get_user_context(db, current_user)
    
    file_context = ""
    if file:
        content = await file.read()
        text_content = extract_text_from_file(content, file.filename)
        if text_content and text_content != "[IMAGE_FILE]":
            file_context = f"\n\nNgười dùng có cung cấp tài liệu đính kèm '{file.filename}' với nội dung:\n{text_content[:10000]}\n(Dựa vào tài liệu này để phân rã nhiệm vụ chính xác hơn nhé)."
        elif text_content == "[IMAGE_FILE]":
            file_context = f"\n\nNgười dùng có đính kèm 1 ảnh '{file.filename}'. Bạn hãy tự suy luận nếu cần thiết."
    elif task_id:
        db_task = db.query(Task).filter(Task.id == task_id, Task.user_id == current_user.id).first()
        if db_task and db_task.attachments:
            try:
                atts = json.loads(db_task.attachments)
                for att in atts:
                    file_path = att.get("path")
                    if attachment_path and file_path != attachment_path:
                        continue
                    if file_path and os.path.exists(file_path):
                        with open(file_path, "rb") as f:
                            content = f.read()
                            text_content = extract_text_from_file(content, att.get("name", ""))
                            if text_content and text_content != "[IMAGE_FILE]":
                                file_context += f"\n\nNội dung file đính kèm '{att.get('name')}':\n{text_content[:10000]}"
            except Exception as e:
                print("Lỗi đọc attachment:", e)

    prompt = f"""You are EduFlow, an expert study planner.
Student info:
{user_ctx}

{LANGUAGE_DETECT_INSTRUCTION}

The user wants to accomplish this goal/task: "{goal}"{file_context}

Break this goal into small, actionable sub-tasks. Estimate the number of Pomodoros for each (1 Pomodoro = 25 minutes).
Sub-task titles MUST be in the same language as the goal or document.

Return ONLY a valid JSON array. No explanation, no markdown code fences.
Example:
[
  {{"title": "Read chapter 1 and take notes", "estimated_pomodoros": 2}},
  {{"title": "Complete exercises at end of chapter 1", "estimated_pomodoros": 3}}
]
"""
    try:
        reply = await call_ai_async(prompt)
        
        # Cắt bỏ markdown nếu AI trả về bất chấp hướng dẫn
        clean_reply = reply.strip()
        if clean_reply.startswith("```json"):
            clean_reply = clean_reply[7:]
        if clean_reply.endswith("```"):
            clean_reply = clean_reply[:-3]
            
        clean_reply = clean_reply.strip()
        data = json.loads(clean_reply)
        
        subtasks = []
        for item in data:
            subtasks.append(SubTaskIdea(
                title=item.get("title", "Nhiệm vụ không tên"),
                estimated_pomodoros=item.get("estimated_pomodoros", 1)
            ))
            
        return BreakdownResponse(subtasks=subtasks)
        
    except json.JSONDecodeError:
        print(f"JSON Parse Error: {reply}")
        raise HTTPException(status_code=500, detail="Phân tích không hợp lệ. Vui lòng thử lại.")
    except Exception as e:
        print(f"AI Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/study-subtask")
async def study_subtask(
    request: StudyRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if not AI_API_KEY:
        raise HTTPException(status_code=500, detail="API Key chưa cấu hình")
        
    task = db.query(Task).filter(Task.id == request.task_id, Task.user_id == current_user.id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhiệm vụ")
    
    # --- Kiểm tra CACHE ---
    if task.ai_recommendation and task.ai_recommendation.strip():
        print(f"[CACHE HIT] Returning saved explanation for task {task.id}")
        return {"explanation": task.ai_recommendation}

    # Xác định context (file đính kèm của nhiệm vụ cha)
    context_text = ""
    target_task = task
    
    # Nếu là nhiệm vụ con, xem file đính kèm của nhiệm vụ cha
    if task.parent_task_id:
        parent = db.query(Task).filter(Task.id == task.parent_task_id).first()
        if parent and parent.attachments:
            try:
                atts = json.loads(parent.attachments)
                if atts:
                    # Thử tìm một tài liệu có thể đọc được
                    for att in atts:
                        file_path = att.get("path")
                        if file_path and os.path.exists(file_path):
                            with open(file_path, "rb") as f:
                                content = f.read()
                                extracted = extract_text_from_file(content, att.get("name", ""))
                                if extracted and extracted != "[IMAGE_FILE]":
                                    context_text += f"\n--- Tài liệu: {att.get('name')} ---\n{extracted[:10000]}\n"
            except:
                pass

    # Xây dựng Prompt
    user_ctx = get_user_context(db, current_user)
    system = f"You are EduFlow AI, an expert tutor. Student info:\n{user_ctx}"
    
    if context_text:
        prompt = f"""{system}

{LANGUAGE_DETECT_INSTRUCTION}
{NO_EMOJI_INSTRUCTION}

Based on the following document:
{context_text}

Create a detailed lesson/explanation about: "{task.title}".
Requirements:
- Clear, concise explanation.
- Highlight key points to remember.
- Provide concrete examples from the document or real-world context.
- Use clean Markdown formatting (headings, lists, bold, tables, code blocks).
- Respond in the same language as the document content.
"""
    else:
        prompt = f"""{system}

{LANGUAGE_DETECT_INSTRUCTION}
{NO_EMOJI_INSTRUCTION}

Create a detailed lesson/explanation about: "{task.title}".
Since there is no attached document, use your expert knowledge to explain as thoroughly as possible.
Requirements:
- Clear, concise explanation.
- Clean Markdown formatting.
- Respond in the same language as the task title.
"""

    try:
        explanation = await call_ai_async(prompt)
        
        # Lưu vào CACHE
        task.ai_recommendation = explanation
        db.commit()
        
        return {"explanation": explanation}
    except Exception as e:
        print(f"AI Error: {e}")
        raise HTTPException(status_code=500, detail="Không thể tạo bài giảng lúc này.")

# ---- AI Practice Exercises ----
@router.post("/practice-subtask")
async def practice_subtask(
    request: StudyRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if not AI_API_KEY:
        raise HTTPException(status_code=500, detail="API Key chưa cấu hình")
        
    task = db.query(Task).filter(Task.id == request.task_id, Task.user_id == current_user.id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhiệm vụ")
    
    # Lấy nội dung bài học 
    lesson_context = task.ai_recommendation or ""
    
    # Kiểm tra file đính kèm 
    doc_context = ""
    if task.parent_task_id:
        parent = db.query(Task).filter(Task.id == task.parent_task_id).first()
        if parent and parent.attachments:
            try:
                atts = json.loads(parent.attachments)
                for att in atts:
                    file_path = att.get("path")
                    if file_path and os.path.exists(file_path):
                        with open(file_path, "rb") as f:
                            content = f.read()
                            extracted = extract_text_from_file(content, att.get("name", ""))
                            if extracted and extracted != "[IMAGE_FILE]":
                                doc_context += f"\n{extracted[:8000]}\n"
            except:
                pass

    user_ctx = get_user_context(db, current_user)
    
    prompt = f"""You are EduFlow AI, an expert tutor creating practice exercises.
Student info:
{user_ctx}

{LANGUAGE_DETECT_INSTRUCTION}
{NO_EMOJI_INSTRUCTION}

Topic: "{task.title}"

{f'Lesson content for reference:{chr(10)}{lesson_context[:6000]}' if lesson_context else ''}
{f'Document context:{chr(10)}{doc_context[:6000]}' if doc_context else ''}

Create a comprehensive practice set for this topic. Include:
1. Multiple choice questions (3-5 questions) with answer options A, B, C, D and the correct answer marked.
2. Short answer questions (2-3 questions) that test understanding.
3. One practical exercise or problem to solve.

Format the output in clean Markdown. At the end, provide an answer key section.
Respond in the same language as the topic/document content.
"""

    try:
        practice = await call_ai_async(prompt)
        return {"practice": practice}
    except Exception as e:
        print(f"AI Error: {e}")
        raise HTTPException(status_code=500, detail="Không thể tạo bài tập lúc này.")

# Task Chat History
@router.get("/task-chat-history")
async def get_task_chat_history(
    task_id: int,
    chat_type: str = "study",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Load lịch sử trò chuyện đã lưu của nhiệm vụ (study hoặc practice)"""
    messages = db.query(TaskChatHistory).filter(
        TaskChatHistory.user_id == current_user.id,
        TaskChatHistory.task_id == task_id,
        TaskChatHistory.chat_type == chat_type
    ).order_by(TaskChatHistory.created_at.asc()).all()
    
    return [{"role": m.role, "message": m.message, "created_at": str(m.created_at)} for m in messages]

# AI Study Chat
@router.post("/study-chat")
async def study_chat(
    request: StudyChatRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if not AI_API_KEY:
        raise HTTPException(status_code=500, detail="API Key chưa cấu hình")
        
    task = db.query(Task).filter(Task.id == request.task_id, Task.user_id == current_user.id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhiệm vụ")
    
    # Lưu tin nhắn user vào DB
    user_msg = TaskChatHistory(
        user_id=current_user.id,
        task_id=request.task_id,
        chat_type=request.chat_type,
        role="user",
        message=request.message
    )
    db.add(user_msg)
    db.commit()
    
    # Lấy nội dung bài học 
    lesson_context = task.ai_recommendation or ""
    
    # Lấy lịch sử trò chuyện gần đây 
    recent_history = db.query(TaskChatHistory).filter(
        TaskChatHistory.user_id == current_user.id,
        TaskChatHistory.task_id == request.task_id,
        TaskChatHistory.chat_type == request.chat_type
    ).order_by(TaskChatHistory.created_at.desc()).limit(10).all()
    
    history_text = ""
    if recent_history:
        for msg in reversed(recent_history):
            role_label = "Student" if msg.role == "user" else "Tutor"
            history_text += f"{role_label}: {msg.message}\n"
    
    # Lấy nội dung tài liệu 
    doc_context = ""
    if task.parent_task_id:
        parent = db.query(Task).filter(Task.id == task.parent_task_id).first()
        if parent and parent.attachments:
            try:
                atts = json.loads(parent.attachments)
                for att in atts:
                    file_path = att.get("path")
                    if file_path and os.path.exists(file_path):
                        with open(file_path, "rb") as f:
                            content = f.read()
                            extracted = extract_text_from_file(content, att.get("name", ""))
                            if extracted and extracted != "[IMAGE_FILE]":
                                doc_context += f"\n{extracted[:6000]}\n"
            except:
                pass

    user_ctx = get_user_context(db, current_user)
    
    prompt = f"""You are EduFlow AI, an expert tutor having a conversation with a student.
Student info:
{user_ctx}

{LANGUAGE_FOLLOW_USER_INSTRUCTION}
{NO_EMOJI_INSTRUCTION}

Context - The student is studying: "{task.title}"

{f'Lesson content already provided:{chr(10)}{lesson_context[:5000]}' if lesson_context else ''}
{f'Source document:{chr(10)}{doc_context[:5000]}' if doc_context else ''}

{f'Recent conversation history:{chr(10)}{history_text}' if history_text else ''}

The student asks: "{request.message}"

Respond helpfully and concisely. Stay focused on the topic. Use Markdown formatting.
IMPORTANT: Respond in the SAME language the student used in their question above.
"""

    try:
        reply = await call_ai_async(prompt, max_tokens=2048)
        
        # Lưu tin nhắn AI vào DB
        ai_msg = TaskChatHistory(
            user_id=current_user.id,
            task_id=request.task_id,
            chat_type=request.chat_type,
            role="ai",
            message=reply
        )
        db.add(ai_msg)
        db.commit()
        
        return {"reply": reply}
    except Exception as e:
        print(f"AI Error: {e}")
        raise HTTPException(status_code=500, detail="Không thể trả lời lúc này.")

# AI Auto-Schedule
@router.post("/auto-schedule", response_model=AutoScheduleResponse)
async def auto_schedule_task(
    request: StudyRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    if not AI_API_KEY:
        raise HTTPException(status_code=500, detail="API Key chưa cấu hình")
    if not request.task_id and not request.subject_id:
        raise HTTPException(status_code=400, detail="Cần cung cấp task_id hoặc subject_id")
        
    duration = 60
    end_limit = datetime.now() + timedelta(days=7)
    task_context = ""
    subject_id_to_save = None
    main_title = ""
    
    if request.task_id:
        task = db.query(Task).filter(Task.id == request.task_id, Task.user_id == current_user.id).first()
        if not task:
            raise HTTPException(status_code=404, detail="Không tìm thấy nhiệm vụ")
            
        subject_id_to_save = task.subject_id
        main_title = f'Nhiệm vụ: {task.title}'
        duration = task.estimated_duration or 60
        if task.due_date:
            end_limit = task.due_date
            
        if getattr(task, "group", None):
            task_context += f"- Belongs to group: {task.group.name}\n"
        if getattr(task, "subject", None):
            task_context += f"- Belongs to subject: {task.subject.subject_name}\n"
            
        subtasks = db.query(Task).filter(Task.parent_task_id == task.id).all()
        if subtasks:
            task_context += "- Subtasks (specific study items to be covered):\n"
            for st in subtasks:
                task_context += f"  * {st.title} (Est. {st.estimated_duration or 25} mins)\n"
    else:
        subject = db.query(Subject).filter(Subject.id == request.subject_id, Subject.user_id == current_user.id).first()
        if not subject:
            raise HTTPException(status_code=404, detail="Không tìm thấy môn học")
            
        subject_id_to_save = subject.id
        main_title = f'Môn học: {subject.subject_name}'
        duration = (subject.target_hours_per_week or 5) * 60
        
        subject_tasks = db.query(Task).filter(Task.subject_id == subject.id).all()
        if subject_tasks:
            task_context += "- Tasks assigned to this subject (specific study items to be covered):\n"
            for st in subject_tasks:
                task_context += f"  * {st.title} (Status: {st.status})\n"

    # Lấy các lịch trình hiện có từ hôm nay đến due_date
    today = datetime.now()
    
    existing_schedules = db.query(Schedule).filter(
        Schedule.user_id == current_user.id,
        Schedule.start_time >= today,
        Schedule.start_time <= end_limit
    ).all()
    
    schedule_text = ""
    for s in existing_schedules:
        schedule_text += f"- {s.start_time.strftime('%Y-%m-%d %H:%M')} -> {s.end_time.strftime('%H:%M')}: {s.title}\n"
        
    user_ctx = get_user_context(db, current_user)
    
    prompt = f"""You are EduFlow AI, an expert study scheduler.
Student info: {user_ctx}

Item to schedule:
- Main Title: "{main_title}"
- Estimated total duration: {duration} minutes
- Deadline for studying: {end_limit.strftime('%Y-%m-%d %H:%M')}
{task_context}

Current scheduled events (DO NOT overlap with these times):
{schedule_text if schedule_text else "No existing schedules found. You have completely free time."}

Rules:
1. Break the total {duration} minutes into manageable study blocks (e.g. 30, 45 or 60 mins each).
2. Distribute them strategically before the deadline. Do NOT place them outside the timeframe between {today.strftime('%Y-%m-%d %H:%M')} and {end_limit.strftime('%Y-%m-%d %H:%M')}.
3. Assign a highly descriptive, practical `title` based on the context. If subtasks are provided, YOU MUST NAME THE SCHEDULE BLOCKS AFTER THE SUBTASKS (e.g. "Học: [Subtask name]"). Do NOT use generic names like "Ôn tập: 1".
4. Set `start_time` and `end_time` strictly in format "YYYY-MM-DDTHH:MM:SS".
5. Provide ONLY a valid JSON array matching exactly this schema, without markdown code fences or explanations:
[
  {{
    "title": "...",
    "description": "...",
    "start_time": "2026-04-03T19:00:00",
    "end_time": "2026-04-03T20:00:00"
  }}
]
"""
    try:
        reply = await call_ai_async(prompt)
        # Cắt bỏ markdown
        clean_reply = reply.strip()
        if clean_reply.startswith("```json"):
            clean_reply = clean_reply[7:]
        elif clean_reply.startswith("```"):
            clean_reply = clean_reply[3:]
        if clean_reply.endswith("```"):
            clean_reply = clean_reply[:-3]
            
        data = json.loads(clean_reply.strip())
        
        blocks = []
        for block in data:
            start_t = datetime.fromisoformat(block['start_time'].replace('Z', ''))
            end_t = datetime.fromisoformat(block['end_time'].replace('Z', ''))
            
            new_schedule = Schedule(
                user_id=current_user.id,
                subject_id=subject_id_to_save,
                title=f"✨ {block.get('title', main_title)}",
                description=block.get('description', ''),
                start_time=start_t,
                end_time=end_t,
                status='scheduled',
                created_by=current_user.id
            )
            db.add(new_schedule)
            db.flush() # flush để lấy ID hoặc tiếp tục
            blocks.append({
                "title": new_schedule.title,
                "description": new_schedule.description,
                "start_time": new_schedule.start_time,
                "end_time": new_schedule.end_time
            })
            
        db.commit()
        invalidate_cache("/stms/schedules", current_user.id)
        
        return {"blocks": blocks}
        
    except json.JSONDecodeError:
        db.rollback()
        print(f"JSON Parse Error: {reply}")
        raise HTTPException(status_code=500, detail="AI trả về phân tích không đúng chuẩn. Vui lòng thử lại.")
    except Exception as e:
        db.rollback()
        print(f"AI Auto-Schedule Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
